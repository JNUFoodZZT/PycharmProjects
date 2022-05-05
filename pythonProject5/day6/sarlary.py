from openpyxl import load_workbook,Workbook
import smtplib
from email.mime.text import MIMEText
from email.header import Header
smtp_obj = smtplib.SMTP_SSL("smtp.exmail.qq.com",465)
smtp_obj.login("6210113257@stu.jiangnan.edu.cn","Zzt20140224")

wb = load_workbook("excel_test.xlsx")
    # 打开excel
sheet = wb['girls']    # 选择序列
for col in sheet.iter_rows(min_row=2,max_row=6,min_col=2):
    row_text = []
    for cell in col:
        row_text.append(cell.value)
    # print(row_text[1])

    mail_body = f'''
    your salary is listed:
    {row_text}
    '''
    msg = MIMEText(mail_body, "html", "utf-8")
    msg["From"] = Header("from jiangnan", "utf-8")
    msg["To"] = Header("hfut", "utf-8")
    msg["Subject"] = Header("testify", 'utf-8')
    smtp_obj.sendmail("6210113257@stu.jiangnan.edu.cn", ["imaphz.qiye.163.com", f"{row_text[1]}"],
                      msg.as_string())
    print(f"successfully,send to {row_text[1]}({row_text[0]})")






