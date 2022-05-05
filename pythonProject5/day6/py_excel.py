from openpyxl import  Workbook, load_workbook
# wb = Workbook()
# ws = wb.active
# print(ws.title)
# ws.title = "girls"
#
# ws["B9"] = "good"
# ws["c9"] = "well"
# ws.append(["ok","no","out"])# 从最下方第一列开始填
# wb.save("excel_test.xlsx")



wb = load_workbook("excel_test.xlsx")
# print(wb.sheetnames)
# print(wb['girls'])
sheet = wb['girls']
print(sheet)
print("-----")
# print((sheet["B2:B5"]))
# print(sheet["B2"].value)

# for row in sheet.iter_rows(min_col = 2,max_col = 3):
#     for cell in row:
#         print(cell.value,end=",")
#     print()           按行循环

for col in sheet.columns:
    for cell in col:
        print(cell.value,end=",")
    print()

sheet["B5"]= "ok111"
wb.save("excel_test.xlsx")

for col in sheet.columns:
    for cell in col:
        print(cell.value,end=",")
    print()
